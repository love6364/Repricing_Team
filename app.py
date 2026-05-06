# # import streamlit as st
# # from openpyxl import load_workbook
# # from openpyxl.utils import get_column_letter
# # import io

# # # ---------------- PAGE CONFIG ----------------
# # st.set_page_config(page_title="Add Price Columns", page_icon="💎", layout="centered")

# # st.title("💎 Add Updated Price & Difference")

# # st.write("Upload your Excel file → Download updated file with formula")

# # # ---------------- FILE UPLOAD ----------------
# # file = st.file_uploader("Upload Excel File", type=["xlsx"])

# # # ---------------- FUNCTION ----------------
# # def process_file(file):
# #     wb = load_workbook(file)
# #     ws = wb.active

# #     # Detect Cost / Cts column
# #     headers = [cell.value for cell in ws[1]]
# #     cost_col = None

# #     for i, h in enumerate(headers):
# #         if h and "cost" in str(h).lower() and "cts" in str(h).lower():
# #             cost_col = i + 1
# #             break

# #     if cost_col is None:
# #         raise Exception("❌ 'Cost / Cts.' column not found")

# #     # Add new columns
# #     last_col = ws.max_column
# #     upd_col = last_col + 1
# #     diff_col = last_col + 2

# #     ws.cell(1, upd_col).value = "Updated Price"
# #     ws.cell(1, diff_col).value = "Difference"

# #     # Column letters
# #     cost_letter = get_column_letter(cost_col)
# #     upd_letter = get_column_letter(upd_col)

# #     # Apply formula
# #     for row in range(2, ws.max_row + 1):
# #         formula = f"=-ROUND(({cost_letter}{row}-{upd_letter}{row})/{cost_letter}{row}*100,2)"
# #         ws.cell(row, diff_col).value = formula

# #     # Save file
# #     output = io.BytesIO()
# #     wb.save(output)
# #     output.seek(0)

# #     return output

# # # ---------------- PROCESS ----------------
# # if file:
# #     try:
# #         output = process_file(file)

# #         st.success("✅ Columns added successfully!")

# #         st.download_button(
# #             label="📥 Download Updated File",
# #             data=output,
# #             file_name="updated_price_file.xlsx"
# #         )

# #     except Exception as e:
# #         st.error(str(e))












# # import streamlit as st
# # from openpyxl import load_workbook
# # from openpyxl.utils import get_column_letter
# # import io

# # # ---------------- PAGE CONFIG ----------------
# # st.set_page_config(page_title="Diamond Smart Repricing", page_icon="💎", layout="centered")

# # st.title("💎 Diamond Smart Repricing Tool")
# # st.write("Upload file → Select multiple shapes → Process → Download")

# # # ---------------- FILE UPLOAD ----------------
# # file = st.file_uploader("Upload Excel File", type=["xlsx"])


# # # ---------------------------------------------------
# # # GET SHAPES FROM FILE
# # # ---------------------------------------------------
# # def get_shapes(uploaded_file):
# #     wb = load_workbook(uploaded_file, data_only=True)
# #     ws = wb.active

# #     headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

# #     shape_col = None

# #     for i, h in enumerate(headers):
# #         if h.lower() == "shape":
# #             shape_col = i + 1
# #             break

# #     if shape_col is None:
# #         return []

# #     shapes = set()

# #     for row in range(2, ws.max_row + 1):
# #         val = ws.cell(row, shape_col).value
# #         if val:
# #             shapes.add(str(val).strip().upper())

# #     return sorted(list(shapes))


# # # ---------------------------------------------------
# # # UI AFTER FILE UPLOAD
# # # ---------------------------------------------------
# # if file:

# #     shape_list = get_shapes(file)

# #     st.subheader("Answer Questions")

# #     selected_shapes = st.multiselect(
# #         "1️⃣ Select Shape(s) you want to work / reprice",
# #         shape_list,
# #         default=shape_list
# #     )

# #     fancy_option = st.radio(
# #         "2️⃣ Do you want to work with Fancy Colors?",
# #         ["Yes", "No"]
# #     )

# #     lot_option = st.radio(
# #         "3️⃣ Remove stones starts with VJ / VP OR ends with A / B / U ?",
# #         ["Yes", "No"]
# #     )

# #     six_cts_option = st.radio(
# #         "4️⃣ Do You want 6.00 <= Stone ?",
# #         ["Yes", "No"]
# #     )

# #     run_btn = st.button("🚀 Process File")


# # # ---------------------------------------------------
# # # PROCESS FILE
# # # ---------------------------------------------------
# #     if run_btn:

# #         try:
# #             wb = load_workbook(file)
# #             ws = wb.active

# #             headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

# #             cost_col = None
# #             color_col = None
# #             shape_col = None
# #             lot_col = None

# #             for i, h in enumerate(headers):
# #                 h_low = h.lower()

# #                 if "cost" in h_low and "cts" in h_low:
# #                     cost_col = i + 1

# #                 elif h_low == "color":
# #                     color_col = i + 1

# #                 elif h_low == "shape":
# #                     shape_col = i + 1

# #                 elif h_low == "lot #":
# #                     lot_col = i + 1

# #             if cost_col is None:
# #                 st.error("❌ Cost / Cts column not found")
# #                 st.stop()

# #             rows_to_delete = []

# #             # --------------------------------
# #             # FILTER ROWS
# #             # --------------------------------
# #             for row in range(2, ws.max_row + 1):

# #                 delete_row = False

# #                 # 1 MULTIPLE SHAPE FILTER
# #                 if shape_col and selected_shapes:
# #                     shp = ws.cell(row, shape_col).value
# #                     shp = str(shp).strip().upper() if shp else ""

# #                     if shp not in selected_shapes:
# #                         delete_row = True

# #                 # 2 FANCY COLOR FILTER
# #                 if not delete_row and fancy_option == "No" and color_col:
# #                     clr = ws.cell(row, color_col).value
# #                     clr = str(clr).strip().upper() if clr else ""

# #                     if clr not in ["D", "E", "F", "G", "H", "I"]:
# #                         delete_row = True

# #                 # 3 LOT FILTER
# #                 if not delete_row and lot_option == "Yes" and lot_col:
# #                     lot = ws.cell(row, lot_col).value
# #                     lot = str(lot).strip().upper() if lot else ""

# #                     if (
# #                         lot.startswith("VJ")
# #                         or lot.startswith("VP")
# #                         or lot.endswith("A")
# #                         or lot.endswith("B")
# #                         or lot.endswith("U")
# #                     ):
# #                         delete_row = True

# #                 # 4 REMOVE 6.00+ STONES
# #                 if not delete_row and six_cts_option == "No":
# #                     val = ws.cell(row, cost_col).value

# #                     try:
# #                         num = float(val)
# #                         if num >= 6.00:
# #                             delete_row = True
# #                     except:
# #                         pass

# #                 if delete_row:
# #                     rows_to_delete.append(row)

# #             # DELETE ROWS REVERSE
# #             for r in reversed(rows_to_delete):
# #                 ws.delete_rows(r)

# #             # --------------------------------
# #             # ADD UPDATED PRICE + DIFFERENCE
# #             # --------------------------------
# #             last_col = ws.max_column
# #             upd_col = last_col + 1
# #             diff_col = last_col + 2

# #             ws.cell(1, upd_col).value = "Updated Price"
# #             ws.cell(1, diff_col).value = "Difference"

# #             cost_letter = get_column_letter(cost_col)
# #             upd_letter = get_column_letter(upd_col)

# #             for row in range(2, ws.max_row + 1):
# #                 formula = f"=-ROUND(({cost_letter}{row}-{upd_letter}{row})/{cost_letter}{row}*100,2)"
# #                 ws.cell(row, diff_col).value = formula

# #             # --------------------------------
# #             # SAVE FILE
# #             # --------------------------------
# #             output = io.BytesIO()
# #             wb.save(output)
# #             output.seek(0)

# #             st.success("✅ File Processed Successfully!")

# #             st.download_button(
# #                 label="📥 Download Final File",
# #                 data=output,
# #                 file_name="diamond_final_output.xlsx"
# #             )

# #         except Exception as e:
# #             st.error(str(e))



# # import streamlit as st
# # import pandas as pd
# # import io
# # import zipfile
# # from openpyxl import load_workbook
# # from openpyxl.utils import get_column_letter

# # # -------------------------------------------------
# # # PAGE CONFIG
# # # -------------------------------------------------
# # st.set_page_config(page_title="Diamond Smart Repricing PRO", page_icon="💎", layout="wide")

# # st.title("💎 Diamond Smart Repricing PRO")
# # st.write("⚡ Fast Version | Upload → Filter → Formula → Download")

# # # -------------------------------------------------
# # # FILE UPLOAD
# # # -------------------------------------------------
# # file = st.file_uploader("Upload Excel File", type=["xlsx"])


# # # -------------------------------------------------
# # # TEAM SHAPE MATCHING
# # # -------------------------------------------------
# # def get_team(shape):
# #     s = str(shape).upper()

# #     if any(x in s for x in ["RBC", "ROUND", "ASSCHER", "PRINCESS"]):
# #         return "Love"

# #     elif any(x in s for x in ["CUSHION MODIFIED", "CUSHION BRILLIANT", "HEART"]):
# #         return "Harshali"

# #     elif any(x in s for x in ["OVAL", "EMERALD"]):
# #         return "Gautam"

# #     elif any(x in s for x in ["PEAR", "RADIANT"]):
# #         return "Milan"

# #     return "Others"


# # # -------------------------------------------------
# # # ADD EXCEL FORMULA
# # # -------------------------------------------------
# # def add_formula_excel(df, cost_col_name):

# #     output = io.BytesIO()

# #     # Save pandas first
# #     with pd.ExcelWriter(output, engine="openpyxl") as writer:
# #         df.to_excel(writer, index=False, sheet_name="Sheet1")

# #     output.seek(0)

# #     # Load with openpyxl
# #     wb = load_workbook(output)
# #     ws = wb.active

# #     headers = [cell.value for cell in ws[1]]

# #     cost_col = headers.index(cost_col_name) + 1
# #     upd_col = len(headers) + 1
# #     diff_col = len(headers) + 2

# #     ws.cell(1, upd_col).value = "Updated Price"
# #     ws.cell(1, diff_col).value = "Difference"

# #     cost_letter = get_column_letter(cost_col)
# #     upd_letter = get_column_letter(upd_col)

# #     for row in range(2, ws.max_row + 1):
# #         ws.cell(row, diff_col).value = f"=-ROUND(({cost_letter}{row}-{upd_letter}{row})/{cost_letter}{row}*100,2)"

# #     final = io.BytesIO()
# #     wb.save(final)
# #     final.seek(0)

# #     return final


# # # -------------------------------------------------
# # # MAIN APP
# # # -------------------------------------------------
# # if file:

# #     with st.spinner("⚡ Reading File..."):
# #         df = pd.read_excel(file)

# #     df.columns = df.columns.str.strip()

# #     # Detect columns
# #     shape_col = "Shape"
# #     color_col = "Color"
# #     lot_col = "Lot #"

# #     cost_col = None
# #     for c in df.columns:
# #         if "cost" in c.lower() and "cts" in c.lower():
# #             cost_col = c
# #             break

# #     if cost_col is None:
# #         st.error("❌ Cost / Cts column not found")
# #         st.stop()

# #     # -------------------------------------------------
# #     # QUESTIONS
# #     # -------------------------------------------------
# #     shapes = sorted(df[shape_col].dropna().astype(str).str.upper().unique())

# #     selected_shapes = st.multiselect(
# #         "1️⃣ Select Shape(s)",
# #         shapes,
# #         default=shapes
# #     )

# #     fancy_option = st.radio(
# #         "2️⃣ Fancy Colors?",
# #         ["Yes", "No"]
# #     )

# #     lot_option = st.radio(
# #         "3️⃣ Remove VP/VJ Starts or Ends A/B ?",
# #         ["Yes", "No"]
# #     )

# #     six_option = st.radio(
# #         "4️⃣ Keep Only 6.00+ Stones?",
# #         ["Yes", "No"]
# #     )

# #     col1, col2 = st.columns(2)

# #     process_btn = col1.button("🚀 Process File")
# #     team_btn = col2.button("👥 Create Team Files")


# # # -------------------------------------------------
# # # PROCESS FILE
# # # -------------------------------------------------
# #     if process_btn:

# #         with st.spinner("⚡ Processing..."):

# #             data = df.copy()

# #             # Shape Filter
# #             data = data[data[shape_col].astype(str).str.upper().isin(selected_shapes)]

# #             # Fancy Color Filter
# #             if fancy_option == "No":
# #                 data = data[data[color_col].astype(str).str.upper().isin(["D","E","F","G","H","I"])]

# #             # Lot Filter
# #             if lot_option == "Yes":
# #                 lot = data[lot_col].astype(str).str.upper()

# #                 data = data[
# #                     ~(
# #                         lot.str.startswith("VP") |
# #                         lot.str.startswith("VJ") |
# #                         lot.str.endswith("A") |
# #                         lot.str.endswith("B")
# #                     )
# #                 ]

# #             # 6.00+
# #             if six_option == "Yes":
# #                 data = data[pd.to_numeric(data[cost_col], errors="coerce") >= 6.00]

# #             final_file = add_formula_excel(data, cost_col)

# #         st.success("✅ Ready!")

# #         st.download_button(
# #             "📥 Download Final File",
# #             data=final_file,
# #             file_name="diamond_output.xlsx"
# #         )


# # # -------------------------------------------------
# # # TEAM FILES
# # # -------------------------------------------------
# #     if team_btn:

# #         with st.spinner("⚡ Creating Team Files..."):

# #             data = df.copy()

# #             data["Team"] = data[shape_col].apply(get_team)

# #             zip_buffer = io.BytesIO()

# #             with zipfile.ZipFile(zip_buffer, "w") as zf:

# #                 for team in ["Love", "Harshali", "Gautam", "Milan"]:

# #                     team_df = data[data["Team"] == team].drop(columns=["Team"])

# #                     if len(team_df) > 0:

# #                         file_data = add_formula_excel(team_df, cost_col)
# #                         zf.writestr(f"{team}.xlsx", file_data.read())

# #             zip_buffer.seek(0)

# #         st.success("✅ Team Files Ready!")

# #         st.download_button(
# #             "📦 Download Team ZIP",
# #             data=zip_buffer,
# #             file_name="Repricing_Team_Files.zip"
# #         )

        


# import streamlit as st
# import pandas as pd
# import io
# import zipfile
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter

# # -------------------------------------------------
# # PAGE CONFIG
# # -------------------------------------------------
# st.set_page_config(page_title="Diamond Smart Repricing PRO", page_icon="💎", layout="wide")

# st.title("💎 Diamond Smart Repricing PRO")
# st.write("⚡ Fast Version | Upload → Filter → Formula → Download")

# # -------------------------------------------------
# # FILE UPLOAD
# # -------------------------------------------------
# file = st.file_uploader("Upload Excel File", type=["xlsx"])

# # -------------------------------------------------
# # TEAM SHAPE MATCHING
# # -------------------------------------------------
# TEAM_SHAPES = {
#     "Love": ["RBC", "ROUND", "ASSCHER", "PRINCESS"],
#     "Harshali": ["CUSHION MODIFIED", "CUSHION BRILLIANT", "HEART"],
#     "Gautam": ["OVAL", "EMERALD"],
#     "Milan": ["PEAR", "RADIANT"]
# }

# def get_team(shape):
#     s = str(shape).upper()

#     for team, shape_list in TEAM_SHAPES.items():
#         if any(x in s for x in shape_list):
#             return team

#     return "Others"

# # -------------------------------------------------
# # ADD EXCEL FORMULA
# # -------------------------------------------------
# def add_formula_excel(df, cost_col_name):

#     output = io.BytesIO()

#     with pd.ExcelWriter(output, engine="openpyxl") as writer:
#         df.to_excel(writer, index=False, sheet_name="Sheet1")

#     output.seek(0)

#     wb = load_workbook(output)
#     ws = wb.active

#     headers = [cell.value for cell in ws[1]]

#     cost_col = headers.index(cost_col_name) + 1
#     upd_col = len(headers) + 1
#     diff_col = len(headers) + 2

#     ws.cell(1, upd_col).value = "Updated Price"
#     ws.cell(1, diff_col).value = "Difference"

#     cost_letter = get_column_letter(cost_col)
#     upd_letter = get_column_letter(upd_col)

#     for row in range(2, ws.max_row + 1):
#         ws.cell(row, diff_col).value = f"=-ROUND(({cost_letter}{row}-{upd_letter}{row})/{cost_letter}{row}*100,2)"

#     final = io.BytesIO()
#     wb.save(final)
#     final.seek(0)

#     return final

# # -------------------------------------------------
# # FILTER FUNCTION
# # -------------------------------------------------
# def apply_filters(data, shapes_selected, fancy_option, lot_option, six_option):

#     data = data[data[shape_col].astype(str).str.upper().isin(shapes_selected)]

#     if fancy_option == "No":
#         data = data[data[color_col].astype(str).str.upper().isin(["D", "E", "F", "G", "H", "I"])]

#     if lot_option == "Yes":
#         lot = data[lot_col].astype(str).str.upper()

#         data = data[
#             ~(
#                 lot.str.startswith("VP") |
#                 lot.str.startswith("VJ") |
#                 lot.str.endswith("A") |
#                 lot.str.endswith("B")
#             )
#         ]

#     if six_option == "Yes":
#         data = data[pd.to_numeric(data[cost_col], errors="coerce") >= 6.00]

#     return data

# # -------------------------------------------------
# # MAIN APP
# # -------------------------------------------------
# if file:

#     df = pd.read_excel(file)
#     df.columns = df.columns.str.strip()

#     shape_col = "Shape"
#     color_col = "Color"
#     lot_col = "Lot #"

#     cost_col = None
#     for c in df.columns:
#         if "cost" in c.lower() and "cts" in c.lower():
#             cost_col = c
#             break

#     if cost_col is None:
#         st.error("Cost/Cts column not found")
#         st.stop()

#     all_shapes = sorted(df[shape_col].dropna().astype(str).str.upper().unique())

#     # -------------------------------------------------
#     # FIRST QUESTION
#     # -------------------------------------------------
#     team_mode = st.radio(
#         "1️⃣ Do you want to create file for Repricing Team?",
#         ["Yes", "No"]
#     )

#     # -------------------------------------------------
#     # SHAPE SELECTION
#     # -------------------------------------------------
#     if team_mode == "No":
#         selected_shapes = st.multiselect(
#             "2️⃣ Select Shape(s)",
#             all_shapes,
#             default=all_shapes
#         )
#     else:
#         selected_shapes = all_shapes

#     # -------------------------------------------------
#     # OTHER QUESTIONS
#     # -------------------------------------------------
#     fancy_option = st.radio("3️⃣ Fancy Colors?", ["Yes", "No"])
#     lot_option = st.radio("4️⃣ Remove VP/VJ Starts or Ends A/B ?", ["Yes", "No"])
#     six_option = st.radio("5️⃣ Keep Only 6.00+ Stones?", ["Yes", "No"])

#     if st.button("🚀 Process & Download"):

#         # -------------------------------------------------
#         # TEAM MODE
#         # -------------------------------------------------
#         if team_mode == "Yes":

#             zip_buffer = io.BytesIO()

#             with zipfile.ZipFile(zip_buffer, "w") as zf:

#                 for team, team_shapes in TEAM_SHAPES.items():

#                     team_df = df.copy()

#                     # shapes assigned automatically
#                     shapes_for_team = [
#                         s for s in all_shapes
#                         if any(x in s for x in team_shapes)
#                     ]

#                     team_df = apply_filters(
#                         team_df,
#                         shapes_for_team,
#                         fancy_option,
#                         lot_option,
#                         six_option
#                     )

#                     if len(team_df) > 0:
#                         file_data = add_formula_excel(team_df, cost_col)
#                         zf.writestr(f"{team}.xlsx", file_data.read())

#             zip_buffer.seek(0)

#             st.download_button(
#                 "📦 Download Team ZIP",
#                 data=zip_buffer,
#                 file_name="Repricing_Team_Files.zip"
#             )

#         # -------------------------------------------------
#         # MANUAL MODE
#         # -------------------------------------------------
#         else:

#             data = df.copy()

#             data = apply_filters(
#                 data,
#                 selected_shapes,
#                 fancy_option,
#                 lot_option,
#                 six_option
#             )

#             final_file = add_formula_excel(data, cost_col)

#             st.download_button(
#                 "📥 Download Final File",
#                 data=final_file,
#                 file_name="diamond_output.xlsx"
#             )

# import streamlit as st
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# import io

# # ---------------- PAGE CONFIG ----------------
# st.set_page_config(page_title="Add Price Columns", page_icon="💎", layout="centered")

# st.title("💎 Add Updated Price & Difference")

# st.write("Upload your Excel file → Download updated file with formula")

# # ---------------- FILE UPLOAD ----------------
# file = st.file_uploader("Upload Excel File", type=["xlsx"])

# # ---------------- FUNCTION ----------------
# def process_file(file):
#     wb = load_workbook(file)
#     ws = wb.active

#     # Detect Cost / Cts column
#     headers = [cell.value for cell in ws[1]]
#     cost_col = None

#     for i, h in enumerate(headers):
#         if h and "cost" in str(h).lower() and "cts" in str(h).lower():
#             cost_col = i + 1
#             break

#     if cost_col is None:
#         raise Exception("❌ 'Cost / Cts.' column not found")

#     # Add new columns
#     last_col = ws.max_column
#     upd_col = last_col + 1
#     diff_col = last_col + 2

#     ws.cell(1, upd_col).value = "Updated Price"
#     ws.cell(1, diff_col).value = "Difference"

#     # Column letters
#     cost_letter = get_column_letter(cost_col)
#     upd_letter = get_column_letter(upd_col)

#     # Apply formula
#     for row in range(2, ws.max_row + 1):
#         formula = f"=-ROUND(({cost_letter}{row}-{upd_letter}{row})/{cost_letter}{row}*100,2)"
#         ws.cell(row, diff_col).value = formula

#     # Save file
#     output = io.BytesIO()
#     wb.save(output)
#     output.seek(0)

#     return output

# # ---------------- PROCESS ----------------
# if file:
#     try:
#         output = process_file(file)

#         st.success("✅ Columns added successfully!")

#         st.download_button(
#             label="📥 Download Updated File",
#             data=output,
#             file_name="updated_price_file.xlsx"
#         )

#     except Exception as e:
#         st.error(str(e))












# import streamlit as st
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# import io

# # ---------------- PAGE CONFIG ----------------
# st.set_page_config(page_title="Diamond Smart Repricing", page_icon="💎", layout="centered")

# st.title("💎 Diamond Smart Repricing Tool")
# st.write("Upload file → Select multiple shapes → Process → Download")

# # ---------------- FILE UPLOAD ----------------
# file = st.file_uploader("Upload Excel File", type=["xlsx"])


# # ---------------------------------------------------
# # GET SHAPES FROM FILE
# # ---------------------------------------------------
# def get_shapes(uploaded_file):
#     wb = load_workbook(uploaded_file, data_only=True)
#     ws = wb.active

#     headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

#     shape_col = None

#     for i, h in enumerate(headers):
#         if h.lower() == "shape":
#             shape_col = i + 1
#             break

#     if shape_col is None:
#         return []

#     shapes = set()

#     for row in range(2, ws.max_row + 1):
#         val = ws.cell(row, shape_col).value
#         if val:
#             shapes.add(str(val).strip().upper())

#     return sorted(list(shapes))


# # ---------------------------------------------------
# # UI AFTER FILE UPLOAD
# # ---------------------------------------------------
# if file:

#     shape_list = get_shapes(file)

#     st.subheader("Answer Questions")

#     selected_shapes = st.multiselect(
#         "1️⃣ Select Shape(s) you want to work / reprice",
#         shape_list,
#         default=shape_list
#     )

#     fancy_option = st.radio(
#         "2️⃣ Do you want to work with Fancy Colors?",
#         ["Yes", "No"]
#     )

#     lot_option = st.radio(
#         "3️⃣ Remove stones starts with VJ / VP OR ends with A / B / U ?",
#         ["Yes", "No"]
#     )

#     six_cts_option = st.radio(
#         "4️⃣ Do You want 6.00 <= Stone ?",
#         ["Yes", "No"]
#     )

#     run_btn = st.button("🚀 Process File")


# # ---------------------------------------------------
# # PROCESS FILE
# # ---------------------------------------------------
#     if run_btn:

#         try:
#             wb = load_workbook(file)
#             ws = wb.active

#             headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

#             cost_col = None
#             color_col = None
#             shape_col = None
#             lot_col = None

#             for i, h in enumerate(headers):
#                 h_low = h.lower()

#                 if "cost" in h_low and "cts" in h_low:
#                     cost_col = i + 1

#                 elif h_low == "color":
#                     color_col = i + 1

#                 elif h_low == "shape":
#                     shape_col = i + 1

#                 elif h_low == "lot #":
#                     lot_col = i + 1

#             if cost_col is None:
#                 st.error("❌ Cost / Cts column not found")
#                 st.stop()

#             rows_to_delete = []

#             # --------------------------------
#             # FILTER ROWS
#             # --------------------------------
#             for row in range(2, ws.max_row + 1):

#                 delete_row = False

#                 # 1 MULTIPLE SHAPE FILTER
#                 if shape_col and selected_shapes:
#                     shp = ws.cell(row, shape_col).value
#                     shp = str(shp).strip().upper() if shp else ""

#                     if shp not in selected_shapes:
#                         delete_row = True

#                 # 2 FANCY COLOR FILTER
#                 if not delete_row and fancy_option == "No" and color_col:
#                     clr = ws.cell(row, color_col).value
#                     clr = str(clr).strip().upper() if clr else ""

#                     if clr not in ["D", "E", "F", "G", "H", "I"]:
#                         delete_row = True

#                 # 3 LOT FILTER
#                 if not delete_row and lot_option == "Yes" and lot_col:
#                     lot = ws.cell(row, lot_col).value
#                     lot = str(lot).strip().upper() if lot else ""

#                     if (
#                         lot.startswith("VJ")
#                         or lot.startswith("VP")
#                         or lot.endswith("A")
#                         or lot.endswith("B")
#                         or lot.endswith("U")
#                     ):
#                         delete_row = True

#                 # 4 REMOVE 6.00+ STONES
#                 if not delete_row and six_cts_option == "No":
#                     val = ws.cell(row, cost_col).value

#                     try:
#                         num = float(val)
#                         if num >= 6.00:
#                             delete_row = True
#                     except:
#                         pass

#                 if delete_row:
#                     rows_to_delete.append(row)

#             # DELETE ROWS REVERSE
#             for r in reversed(rows_to_delete):
#                 ws.delete_rows(r)

#             # --------------------------------
#             # ADD UPDATED PRICE + DIFFERENCE
#             # --------------------------------
#             last_col = ws.max_column
#             upd_col = last_col + 1
#             diff_col = last_col + 2

#             ws.cell(1, upd_col).value = "Updated Price"
#             ws.cell(1, diff_col).value = "Difference"

#             cost_letter = get_column_letter(cost_col)
#             upd_letter = get_column_letter(upd_col)

#             for row in range(2, ws.max_row + 1):
#                 formula = f"=-ROUND(({cost_letter}{row}-{upd_letter}{row})/{cost_letter}{row}*100,2)"
#                 ws.cell(row, diff_col).value = formula

#             # --------------------------------
#             # SAVE FILE
#             # --------------------------------
#             output = io.BytesIO()
#             wb.save(output)
#             output.seek(0)

#             st.success("✅ File Processed Successfully!")

#             st.download_button(
#                 label="📥 Download Final File",
#                 data=output,
#                 file_name="diamond_final_output.xlsx"
#             )

#         except Exception as e:
#             st.error(str(e))



# import streamlit as st
# import pandas as pd
# import io
# import zipfile
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter

# # -------------------------------------------------
# # PAGE CONFIG
# # -------------------------------------------------
# st.set_page_config(page_title="Diamond Smart Repricing PRO", page_icon="💎", layout="wide")

# st.title("💎 Diamond Smart Repricing PRO")
# st.write("⚡ Fast Version | Upload → Filter → Formula → Download")

# # -------------------------------------------------
# # FILE UPLOAD
# # -------------------------------------------------
# file = st.file_uploader("Upload Excel File", type=["xlsx"])


# # -------------------------------------------------
# # TEAM SHAPE MATCHING
# # -------------------------------------------------
# def get_team(shape):
#     s = str(shape).upper()

#     if any(x in s for x in ["RBC", "ROUND", "ASSCHER", "PRINCESS"]):
#         return "Love"

#     elif any(x in s for x in ["CUSHION MODIFIED", "CUSHION BRILLIANT", "HEART"]):
#         return "Harshali"

#     elif any(x in s for x in ["OVAL", "EMERALD"]):
#         return "Gautam"

#     elif any(x in s for x in ["PEAR", "RADIANT"]):
#         return "Milan"

#     return "Others"


# # -------------------------------------------------
# # ADD EXCEL FORMULA
# # -------------------------------------------------
# def add_formula_excel(df, cost_col_name):

#     output = io.BytesIO()

#     # Save pandas first
#     with pd.ExcelWriter(output, engine="openpyxl") as writer:
#         df.to_excel(writer, index=False, sheet_name="Sheet1")

#     output.seek(0)

#     # Load with openpyxl
#     wb = load_workbook(output)
#     ws = wb.active

#     headers = [cell.value for cell in ws[1]]

#     cost_col = headers.index(cost_col_name) + 1
#     upd_col = len(headers) + 1
#     diff_col = len(headers) + 2

#     ws.cell(1, upd_col).value = "Updated Price"
#     ws.cell(1, diff_col).value = "Difference"

#     cost_letter = get_column_letter(cost_col)
#     upd_letter = get_column_letter(upd_col)

#     for row in range(2, ws.max_row + 1):
#         ws.cell(row, diff_col).value = f"=-ROUND(({cost_letter}{row}-{upd_letter}{row})/{cost_letter}{row}*100,2)"

#     final = io.BytesIO()
#     wb.save(final)
#     final.seek(0)

#     return final


# # -------------------------------------------------
# # MAIN APP
# # -------------------------------------------------
# if file:

#     with st.spinner("⚡ Reading File..."):
#         df = pd.read_excel(file)

#     df.columns = df.columns.str.strip()

#     # Detect columns
#     shape_col = "Shape"
#     color_col = "Color"
#     lot_col = "Lot #"

#     cost_col = None
#     for c in df.columns:
#         if "cost" in c.lower() and "cts" in c.lower():
#             cost_col = c
#             break

#     if cost_col is None:
#         st.error("❌ Cost / Cts column not found")
#         st.stop()

#     # -------------------------------------------------
#     # QUESTIONS
#     # -------------------------------------------------
#     shapes = sorted(df[shape_col].dropna().astype(str).str.upper().unique())

#     selected_shapes = st.multiselect(
#         "1️⃣ Select Shape(s)",
#         shapes,
#         default=shapes
#     )

#     fancy_option = st.radio(
#         "2️⃣ Fancy Colors?",
#         ["Yes", "No"]
#     )

#     lot_option = st.radio(
#         "3️⃣ Remove VP/VJ Starts or Ends A/B ?",
#         ["Yes", "No"]
#     )

#     six_option = st.radio(
#         "4️⃣ Keep Only 6.00+ Stones?",
#         ["Yes", "No"]
#     )

#     col1, col2 = st.columns(2)

#     process_btn = col1.button("🚀 Process File")
#     team_btn = col2.button("👥 Create Team Files")


# # -------------------------------------------------
# # PROCESS FILE
# # -------------------------------------------------
#     if process_btn:

#         with st.spinner("⚡ Processing..."):

#             data = df.copy()

#             # Shape Filter
#             data = data[data[shape_col].astype(str).str.upper().isin(selected_shapes)]

#             # Fancy Color Filter
#             if fancy_option == "No":
#                 data = data[data[color_col].astype(str).str.upper().isin(["D","E","F","G","H","I"])]

#             # Lot Filter
#             if lot_option == "Yes":
#                 lot = data[lot_col].astype(str).str.upper()

#                 data = data[
#                     ~(
#                         lot.str.startswith("VP") |
#                         lot.str.startswith("VJ") |
#                         lot.str.endswith("A") |
#                         lot.str.endswith("B")
#                     )
#                 ]

#             # 6.00+
#             if six_option == "Yes":
#                 data = data[pd.to_numeric(data[cost_col], errors="coerce") >= 6.00]

#             final_file = add_formula_excel(data, cost_col)

#         st.success("✅ Ready!")

#         st.download_button(
#             "📥 Download Final File",
#             data=final_file,
#             file_name="diamond_output.xlsx"
#         )


# # -------------------------------------------------
# # TEAM FILES
# # -------------------------------------------------
#     if team_btn:

#         with st.spinner("⚡ Creating Team Files..."):

#             data = df.copy()

#             data["Team"] = data[shape_col].apply(get_team)

#             zip_buffer = io.BytesIO()

#             with zipfile.ZipFile(zip_buffer, "w") as zf:

#                 for team in ["Love", "Harshali", "Gautam", "Milan"]:

#                     team_df = data[data["Team"] == team].drop(columns=["Team"])

#                     if len(team_df) > 0:

#                         file_data = add_formula_excel(team_df, cost_col)
#                         zf.writestr(f"{team}.xlsx", file_data.read())

#             zip_buffer.seek(0)

#         st.success("✅ Team Files Ready!")

#         st.download_button(
#             "📦 Download Team ZIP",
#             data=zip_buffer,
#             file_name="Repricing_Team_Files.zip"
#         )

        


# import streamlit as st
# import pandas as pd
# import io
# import zipfile
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter

# # -------------------------------------------------
# # PAGE CONFIG
# # -------------------------------------------------
# st.set_page_config(page_title="Diamond Smart Repricing PRO", page_icon="💎", layout="wide")

# st.title("💎 Diamond Smart Repricing PRO")
# st.write("⚡ Fast Version | Upload → Filter → Formula → Download")

# # -------------------------------------------------
# # FILE UPLOAD
# # -------------------------------------------------
# file = st.file_uploader("Upload Excel File", type=["xlsx"])

# # -------------------------------------------------
# # TEAM SHAPE MATCHING
# # -------------------------------------------------
# TEAM_SHAPES = {
#     "Love": ["RBC", "ROUND", "ASSCHER", "PRINCESS"],
#     "Harshali": ["CUSHION MODIFIED", "CUSHION BRILLIANT", "HEART"],
#     "Gautam": ["OVAL", "EMERALD"],
#     "Milan": ["PEAR", "RADIANT"]
# }

# def get_team(shape):
#     s = str(shape).upper()

#     for team, shape_list in TEAM_SHAPES.items():
#         if any(x in s for x in shape_list):
#             return team

#     return "Others"

# # -------------------------------------------------
# # ADD EXCEL FORMULA
# # -------------------------------------------------
# def add_formula_excel(df, cost_col_name):

#     output = io.BytesIO()

#     with pd.ExcelWriter(output, engine="openpyxl") as writer:
#         df.to_excel(writer, index=False, sheet_name="Sheet1")

#     output.seek(0)

#     wb = load_workbook(output)
#     ws = wb.active

#     headers = [cell.value for cell in ws[1]]

#     cost_col = headers.index(cost_col_name) + 1
#     upd_col = len(headers) + 1
#     diff_col = len(headers) + 2

#     ws.cell(1, upd_col).value = "Updated Price"
#     ws.cell(1, diff_col).value = "Difference"

#     cost_letter = get_column_letter(cost_col)
#     upd_letter = get_column_letter(upd_col)

#     for row in range(2, ws.max_row + 1):
#         ws.cell(row, diff_col).value = f"=-ROUND(({cost_letter}{row}-{upd_letter}{row})/{cost_letter}{row}*100,2)"

#     final = io.BytesIO()
#     wb.save(final)
#     final.seek(0)

#     return final

# # -------------------------------------------------
# # FILTER FUNCTION
# # -------------------------------------------------
# def apply_filters(data, shapes_selected, fancy_option, lot_option, six_option):

#     data = data[data[shape_col].astype(str).str.upper().isin(shapes_selected)]

#     if fancy_option == "No":
#         data = data[data[color_col].astype(str).str.upper().isin(["D", "E", "F", "G", "H", "I"])]

#     if lot_option == "Yes":
#         lot = data[lot_col].astype(str).str.upper()

#         data = data[
#             ~(
#                 lot.str.startswith("VP") |
#                 lot.str.startswith("VJ") |
#                 lot.str.endswith("A") |
#                 lot.str.endswith("B")
#             )
#         ]

#     if six_option == "Yes":
#         data = data[pd.to_numeric(data[cost_col], errors="coerce") >= 6.00]

#     return data

# # -------------------------------------------------
# # MAIN APP
# # -------------------------------------------------
# if file:

#     df = pd.read_excel(file)
#     df.columns = df.columns.str.strip()

#     shape_col = "Shape"
#     color_col = "Color"
#     lot_col = "Lot #"

#     cost_col = None
#     for c in df.columns:
#         if "cost" in c.lower() and "cts" in c.lower():
#             cost_col = c
#             break

#     if cost_col is None:
#         st.error("Cost/Cts column not found")
#         st.stop()

#     all_shapes = sorted(df[shape_col].dropna().astype(str).str.upper().unique())

#     # -------------------------------------------------
#     # FIRST QUESTION
#     # -------------------------------------------------
#     team_mode = st.radio(
#         "1️⃣ Do you want to create file for Repricing Team?",
#         ["Yes", "No"]
#     )

#     # -------------------------------------------------
#     # SHAPE SELECTION
#     # -------------------------------------------------
#     if team_mode == "No":
#         selected_shapes = st.multiselect(
#             "2️⃣ Select Shape(s)",
#             all_shapes,
#             default=all_shapes
#         )
#     else:
#         selected_shapes = all_shapes

#     # -------------------------------------------------
#     # OTHER QUESTIONS
#     # -------------------------------------------------
#     fancy_option = st.radio("3️⃣ Fancy Colors?", ["Yes", "No"])
#     lot_option = st.radio("4️⃣ Remove VP/VJ Starts or Ends A/B ?", ["Yes", "No"])
#     six_option = st.radio("5️⃣ Keep Only 6.00+ Stones?", ["Yes", "No"])

#     if st.button("🚀 Process & Download"):

#         # -------------------------------------------------
#         # TEAM MODE
#         # -------------------------------------------------
#         if team_mode == "Yes":

#             zip_buffer = io.BytesIO()

#             with zipfile.ZipFile(zip_buffer, "w") as zf:

#                 for team, team_shapes in TEAM_SHAPES.items():

#                     team_df = df.copy()

#                     # shapes assigned automatically
#                     shapes_for_team = [
#                         s for s in all_shapes
#                         if any(x in s for x in team_shapes)
#                     ]

#                     team_df = apply_filters(
#                         team_df,
#                         shapes_for_team,
#                         fancy_option,
#                         lot_option,
#                         six_option
#                     )

#                     if len(team_df) > 0:
#                         file_data = add_formula_excel(team_df, cost_col)
#                         zf.writestr(f"{team}.xlsx", file_data.read())

#             zip_buffer.seek(0)

#             st.download_button(
#                 "📦 Download Team ZIP",
#                 data=zip_buffer,
#                 file_name="Repricing_Team_Files.zip"
#             )

#         # -------------------------------------------------
#         # MANUAL MODE
#         # -------------------------------------------------
#         else:

#             data = df.copy()

#             data = apply_filters(
#                 data,
#                 selected_shapes,
#                 fancy_option,
#                 lot_option,
#                 six_option
#             )

#             final_file = add_formula_excel(data, cost_col)

#             st.download_button(
#                 "📥 Download Final File",
#                 data=final_file,
#                 file_name="diamond_output.xlsx"
#             )


# import streamlit as st
# import pandas as pd
# import io
# import zipfile
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter

# # -------------------------------------------------
# # PAGE CONFIG
# # -------------------------------------------------
# st.set_page_config(page_title="Diamond Smart Repricing PRO", page_icon="💎", layout="wide")

# st.title("💎 Diamond Smart Repricing PRO")
# st.write("⚡ Fast Version | Upload → Filter → Formula → Download")

# # -------------------------------------------------
# # FILE UPLOAD
# # -------------------------------------------------
# file = st.file_uploader("Upload Excel File", type=["xlsx"])

# # -------------------------------------------------
# # TEAM SHAPE MATCHING
# # -------------------------------------------------
# TEAM_SHAPES = {
#     "Love": ["RBC", "ROUND", "ASSCHER", "PRINCESS"],
#     "Harshali": ["CUSHION MODIFIED", "CUSHION BRILLIANT", "HEART"],
#     "Gautam": ["OVAL", "EMERALD"],
#     "Milan": ["PEAR", "RADIANT"]
# }

# def get_team(shape):
#     s = str(shape).upper()

#     for team, shape_list in TEAM_SHAPES.items():
#         if any(x in s for x in shape_list):
#             return team

#     return "Others"

# # -------------------------------------------------
# # ADD EXCEL FORMULA
# # -------------------------------------------------
# def add_formula_excel(df, cost_col_name):

#     output = io.BytesIO()

#     with pd.ExcelWriter(output, engine="openpyxl") as writer:
#         df.to_excel(writer, index=False, sheet_name="Sheet1")

#     output.seek(0)

#     wb = load_workbook(output)
#     ws = wb.active

#     headers = [cell.value for cell in ws[1]]

#     cost_col = headers.index(cost_col_name) + 1
#     upd_col = len(headers) + 1
#     diff_col = len(headers) + 2

#     ws.cell(1, upd_col).value = "Updated Price"
#     ws.cell(1, diff_col).value = "Difference"

#     cost_letter = get_column_letter(cost_col)
#     upd_letter = get_column_letter(upd_col)

#     for row in range(2, ws.max_row + 1):
#         ws.cell(row, diff_col).value = f"=-ROUND(({cost_letter}{row}-{upd_letter}{row})/{cost_letter}{row}*100,2)"

#     final = io.BytesIO()
#     wb.save(final)
#     final.seek(0)

#     return final

# # -------------------------------------------------
# # FILTER FUNCTION
# # -------------------------------------------------
# def apply_filters(data, shapes_selected, fancy_option, lot_option, six_option):

#     data = data[data[shape_col].astype(str).str.upper().isin(shapes_selected)]

#     if fancy_option == "No":
#         data = data[data[color_col].astype(str).str.upper().isin(["D", "E", "F", "G", "H", "I"])]

#     if lot_option == "Yes":
#         lot = data[lot_col].astype(str).str.upper()

#         data = data[
#             ~(
#                 lot.str.startswith("VP") |
#                 lot.str.startswith("VJ") |
#                 lot.str.endswith("A") |
#                 lot.str.endswith("B")
#             )
#         ]

#     if six_option == "Yes":
#         data = data[pd.to_numeric(data[cost_col], errors="coerce") >= 6.00]

#     return data

# # -------------------------------------------------
# # MAIN APP
# # -------------------------------------------------
# if file:

#     df = pd.read_excel(file)
#     df.columns = df.columns.str.strip()

#     shape_col = "Shape"
#     color_col = "Color"
#     lot_col = "Lot #"

#     cost_col = None
#     for c in df.columns:
#         if "cost" in c.lower() and "cts" in c.lower():
#             cost_col = c
#             break

#     if cost_col is None:
#         st.error("Cost/Cts column not found")
#         st.stop()

#     all_shapes = sorted(df[shape_col].dropna().astype(str).str.upper().unique())

#     # -------------------------------------------------
#     # FIRST QUESTION
#     # -------------------------------------------------
#     team_mode = st.radio(
#         "1️⃣ Do you want to create file for Repricing Team?",
#         ["Yes", "No"]
#     )

#     # -------------------------------------------------
#     # SHAPE SELECTION
#     # -------------------------------------------------
#     if team_mode == "No":
#         selected_shapes = st.multiselect(
#             "2️⃣ Select Shape(s)",
#             all_shapes,
#             default=all_shapes
#         )
#     else:
#         selected_shapes = all_shapes

#     # -------------------------------------------------
#     # OTHER QUESTIONS
#     # -------------------------------------------------
#     fancy_option = st.radio("3️⃣ Fancy Colors?", ["Yes", "No"])
#     lot_option = st.radio("4️⃣ Remove VP/VJ Starts or Ends A/B ?", ["Yes", "No"])
#     six_option = st.radio("5️⃣ Keep Only 6.00+ Stones?", ["Yes", "No"])

#     if st.button("🚀 Process & Download"):

#         # -------------------------------------------------
#         # TEAM MODE
#         # -------------------------------------------------
#         if team_mode == "Yes":

#             zip_buffer = io.BytesIO()

#             with zipfile.ZipFile(zip_buffer, "w") as zf:

#                 for team, team_shapes in TEAM_SHAPES.items():

#                     team_df = df.copy()

#                     # shapes assigned automatically
#                     shapes_for_team = [
#                         s for s in all_shapes
#                         if any(x in s for x in team_shapes)
#                     ]

#                     team_df = apply_filters(
#                         team_df,
#                         shapes_for_team,
#                         fancy_option,
#                         lot_option,
#                         six_option
#                     )

#                     if len(team_df) > 0:
#                         file_data = add_formula_excel(team_df, cost_col)
#                         zf.writestr(f"{team}.xlsx", file_data.read())

#             zip_buffer.seek(0)

#             st.download_button(
#                 "📦 Download Team ZIP",
#                 data=zip_buffer,
#                 file_name="Repricing_Team_Files.zip"
#             )

#         # -------------------------------------------------
#         # MANUAL MODE
#         # -------------------------------------------------
#         else:

#             data = df.copy()

#             data = apply_filters(
#                 data,
#                 selected_shapes,
#                 fancy_option,
#                 lot_option,
#                 six_option
#             )

#             final_file = add_formula_excel(data, cost_col)

#             st.download_button(
#                 "📥 Download Final File",
#                 data=final_file,
#                 file_name="diamond_output.xlsx"
#             )


# import streamlit as st
# import pandas as pd
# import io
# import zipfile
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter

# # -------------------------------------------------
# # PAGE CONFIG
# # -------------------------------------------------
# st.set_page_config(
#     page_title="Diamond Repricing Studio",
#     page_icon="💎",
#     layout="wide"
# )

# # -------------------------------------------------
# # LUXURY CSS
# # -------------------------------------------------
# st.markdown("""
# <style>
# @import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@300;400;500;600&family=Jost:wght@300;400;500;600&display=swap');

# /* ── Base ── */
# html, body, [class*="css"] {
#     font-family: 'Jost', sans-serif;
#     background: #080c14;
#     color: #d6c9b0;
# }
# .stApp { background: #080c14; }
# header[data-testid="stHeader"] { background: transparent; }
# section[data-testid="stSidebar"] { display: none; }

# /* ── Hero ── */
# .hero-wrap {
#     text-align: center;
#     padding: 3.2rem 0 2rem 0;
#     position: relative;
# }
# .hero-gem {
#     font-size: 2.8rem;
#     margin-bottom: 0.5rem;
#     display: block;
#     animation: floatgem 3s ease-in-out infinite;
# }
# @keyframes floatgem {
#     0%,100% { transform: translateY(0); }
#     50%      { transform: translateY(-6px); }
# }
# .hero-title {
#     font-family: 'Cormorant Garamond', serif;
#     font-size: 3rem;
#     font-weight: 400;
#     letter-spacing: 2px;
#     color: #f0e6cc;
#     margin: 0 0 0.4rem 0;
#     line-height: 1.1;
# }
# .hero-sub {
#     font-size: 0.82rem;
#     font-weight: 300;
#     letter-spacing: 3px;
#     text-transform: uppercase;
#     color: #c9a84c;
#     margin-bottom: 0;
# }
# .gold-line {
#     width: 60px;
#     height: 1px;
#     background: linear-gradient(90deg, transparent, #c9a84c, transparent);
#     margin: 1.2rem auto 0 auto;
# }

# /* ── Upload zone ── */
# .upload-section {
#     background: #0d1220;
#     border: 1px solid #1e2535;
#     border-radius: 16px;
#     padding: 2rem;
#     margin-bottom: 1.5rem;
#     position: relative;
#     overflow: hidden;
# }
# .upload-section::before {
#     content: '';
#     position: absolute;
#     top: 0; left: 0; right: 0;
#     height: 2px;
#     background: linear-gradient(90deg, transparent, #c9a84c 50%, transparent);
# }
# .section-label {
#     font-family: 'Cormorant Garamond', serif;
#     font-size: 1.4rem;
#     font-weight: 400;
#     color: #f0e6cc;
#     margin-bottom: 0.2rem;
# }
# .section-desc {
#     font-size: 0.78rem;
#     font-weight: 300;
#     letter-spacing: 1.5px;
#     text-transform: uppercase;
#     color: #c9a84c;
#     margin-bottom: 1rem;
# }

# /* ── Question cards ── */
# .q-card {
#     background: #0d1220;
#     border: 1px solid #1e2535;
#     border-radius: 12px;
#     padding: 1.4rem 1.6rem;
#     margin-bottom: 1rem;
#     transition: border-color 0.25s;
#     position: relative;
# }
# .q-card:hover { border-color: #c9a84c44; }
# .q-num {
#     font-size: 0.65rem;
#     letter-spacing: 2px;
#     text-transform: uppercase;
#     color: #c9a84c;
#     font-weight: 500;
#     margin-bottom: 0.35rem;
# }
# .q-title {
#     font-family: 'Cormorant Garamond', serif;
#     font-size: 1.15rem;
#     font-weight: 500;
#     color: #f0e6cc;
#     margin-bottom: 0.15rem;
# }
# .q-hint {
#     font-size: 0.76rem;
#     color: #6b7280;
#     font-weight: 300;
#     margin-bottom: 0.8rem;
# }

# /* ── Streamlit radio overrides ── */
# div[data-testid="stRadio"] label {
#     font-family: 'Jost', sans-serif !important;
#     font-size: 0.9rem !important;
#     color: #d6c9b0 !important;
#     font-weight: 400 !important;
# }
# div[data-testid="stRadio"] > div { gap: 0.5rem; }

# /* ── Multiselect ── */
# div[data-testid="stMultiSelect"] span {
#     background: #1a2235 !important;
#     color: #d6c9b0 !important;
#     border: 1px solid #2a3550 !important;
#     border-radius: 6px !important;
# }
# div[data-testid="stMultiSelect"] [data-baseweb="tag"] {
#     background: #1f2d45 !important;
# }

# /* ── File uploader ── */
# div[data-testid="stFileUploader"] section {
#     background: #0a0f1a !important;
#     border: 1.5px dashed #2a3550 !important;
#     border-radius: 12px !important;
#     transition: border-color 0.2s;
# }
# div[data-testid="stFileUploader"] section:hover {
#     border-color: #c9a84c !important;
# }
# div[data-testid="stFileUploader"] p { color: #6b7280 !important; }

# /* ── Process button ── */
# div[data-testid="stButton"] > button {
#     background: linear-gradient(135deg, #c9a84c 0%, #a8833a 100%);
#     color: #080c14;
#     border: none;
#     border-radius: 10px;
#     padding: 0.85rem 2rem;
#     font-family: 'Jost', sans-serif;
#     font-size: 0.9rem;
#     font-weight: 600;
#     letter-spacing: 2px;
#     text-transform: uppercase;
#     width: 100%;
#     cursor: pointer;
#     box-shadow: 0 4px 24px rgba(201, 168, 76, 0.25);
#     transition: opacity 0.2s, transform 0.15s, box-shadow 0.2s;
# }
# div[data-testid="stButton"] > button:hover {
#     opacity: 0.92;
#     transform: translateY(-2px);
#     box-shadow: 0 8px 30px rgba(201, 168, 76, 0.35);
# }
# div[data-testid="stButton"] > button:disabled {
#     background: #1e2535 !important;
#     color: #3a4560 !important;
#     box-shadow: none !important;
#     transform: none !important;
# }

# /* ── Download buttons ── */
# div[data-testid="stDownloadButton"] > button {
#     background: #0d1220;
#     color: #c9a84c;
#     border: 1px solid #c9a84c55;
#     border-radius: 10px;
#     padding: 0.7rem 1.4rem;
#     font-family: 'Jost', sans-serif;
#     font-size: 0.85rem;
#     font-weight: 500;
#     letter-spacing: 1.5px;
#     text-transform: uppercase;
#     width: 100%;
#     transition: background 0.2s, border-color 0.2s;
# }
# div[data-testid="stDownloadButton"] > button:hover {
#     background: #c9a84c15;
#     border-color: #c9a84c;
# }

# /* ── Info / success alerts ── */
# div[data-testid="stAlert"] {
#     border-radius: 10px;
#     font-family: 'Jost', sans-serif;
#     font-size: 0.88rem;
# }

# /* ── Divider ── */
# .gold-divider {
#     height: 1px;
#     background: linear-gradient(90deg, transparent, #c9a84c44, transparent);
#     margin: 2rem 0;
# }

# /* ── Team badges ── */
# .team-grid {
#     display: flex; flex-wrap: wrap; gap: 0.7rem; margin-top: 0.5rem;
# }
# .team-badge {
#     background: #111827;
#     border: 1px solid #2a3550;
#     border-radius: 8px;
#     padding: 0.4rem 0.9rem;
#     font-size: 0.78rem;
#     color: #d6c9b0;
#     letter-spacing: 0.5px;
# }
# .team-badge strong { color: #c9a84c; }

# /* ── Result banner ── */
# .result-banner {
#     background: linear-gradient(135deg, #0d1a10, #0d1220);
#     border: 1px solid #1a5c2a;
#     border-radius: 12px;
#     padding: 1.2rem 1.6rem;
#     margin-top: 1rem;
#     display: flex;
#     align-items: center;
#     gap: 1rem;
# }
# .result-icon { font-size: 1.6rem; }
# .result-text { font-size: 0.9rem; color: #6fcf97; }
# .result-text strong { color: #a8e6bf; font-weight: 600; }
# </style>
# """, unsafe_allow_html=True)

# # -------------------------------------------------
# # TEAM CONFIG
# # -------------------------------------------------
# TEAM_SHAPES = {
#     "Love":    ["RBC", "ROUND", "ASSCHER", "PRINCESS"],
#     "Harshali":["CUSHION MODIFIED", "CUSHION BRILLIANT", "HEART"],
#     "Gautam":  ["OVAL", "EMERALD"],
#     "Milan":   ["PEAR", "RADIANT"]
# }

# TEAM_ICONS = {"Love": "🌸", "Harshali": "💫", "Gautam": "🌿", "Milan": "✦"}

# # -------------------------------------------------
# # HELPERS
# # -------------------------------------------------
# def get_team(shape):
#     s = str(shape).upper()
#     for team, shapes in TEAM_SHAPES.items():
#         if any(x in s for x in shapes):
#             return team
#     return "Others"

# def add_formula_excel(df, cost_col_name):
#     output = io.BytesIO()
#     with pd.ExcelWriter(output, engine="openpyxl") as writer:
#         df.to_excel(writer, index=False, sheet_name="Sheet1")
#     output.seek(0)
#     wb = load_workbook(output)
#     ws = wb.active
#     headers = [cell.value for cell in ws[1]]
#     cost_col = headers.index(cost_col_name) + 1
#     upd_col  = len(headers) + 1
#     diff_col = len(headers) + 2
#     ws.cell(1, upd_col).value  = "Updated Price"
#     ws.cell(1, diff_col).value = "Difference"
#     cost_letter = get_column_letter(cost_col)
#     upd_letter  = get_column_letter(upd_col)
#     for row in range(2, ws.max_row + 1):
#         ws.cell(row, diff_col).value = f"=-ROUND(({cost_letter}{row}-{upd_letter}{row})/{cost_letter}{row}*100,2)"
#     final = io.BytesIO()
#     wb.save(final)
#     final.seek(0)
#     return final

# def apply_filters(data, shapes_selected, fancy_option, lot_option, six_option):
#     data = data[data[shape_col].astype(str).str.upper().isin(shapes_selected)]
#     if fancy_option == "Exclude fancy colors (D–I only)":
#         data = data[data[color_col].astype(str).str.upper().isin(["D","E","F","G","H","I"])]
#     if lot_option == "Remove flagged lots":
#         lot = data[lot_col].astype(str).str.upper()
#         data = data[~(lot.str.startswith("VP") | lot.str.startswith("VJ") | lot.str.endswith("A") | lot.str.endswith("B"))]
#     if six_option == "6.00 and above only":
#         data = data[pd.to_numeric(data[cost_col], errors="coerce") >= 6.00]
#     return data

# # -------------------------------------------------
# # HERO
# # -------------------------------------------------
# st.markdown("""
# <div class="hero-wrap">
#     <span class="hero-gem">💎</span>
#     <h1 class="hero-title">Diamond Repricing Studio</h1>
#     <p class="hero-sub">Smart Repricing · Team Distribution · Formula Ready</p>
#     <div class="gold-line"></div>
# </div>
# """, unsafe_allow_html=True)

# # -------------------------------------------------
# # FILE UPLOAD
# # -------------------------------------------------
# st.markdown("""
# <div class="upload-section">
#     <div class="section-label">Source File</div>
#     <div class="section-desc">Upload your master diamond inventory</div>
# """, unsafe_allow_html=True)

# file = st.file_uploader(
#     "Upload Excel File",
#     type=["xlsx"],
#     label_visibility="collapsed"
# )

# st.markdown("</div>", unsafe_allow_html=True)

# # -------------------------------------------------
# # MAIN FLOW
# # -------------------------------------------------
# if file:
#     df = pd.read_excel(file)
#     df.columns = df.columns.str.strip()

#     shape_col = "Shape"
#     color_col = "Color"
#     lot_col   = "Lot #"

#     cost_col = None
#     for c in df.columns:
#         if "cost" in c.lower() and "cts" in c.lower():
#             cost_col = c
#             break

#     if cost_col is None:
#         st.error("⚠️ Could not locate a Cost/Cts column. Please verify your file headers.")
#         st.stop()

#     all_shapes = sorted(df[shape_col].dropna().astype(str).str.upper().unique())

#     # File summary pill
#     st.markdown(f"""
#     <div style="display:flex; gap:1rem; margin-bottom:1.5rem; flex-wrap:wrap;">
#         <div style="background:#0d1220; border:1px solid #1e2535; border-radius:8px; padding:0.5rem 1rem; font-size:0.8rem;">
#             <span style="color:#c9a84c; letter-spacing:1px; text-transform:uppercase; font-size:0.65rem;">Stones</span><br>
#             <span style="color:#f0e6cc; font-weight:500;">{len(df):,}</span>
#         </div>
#         <div style="background:#0d1220; border:1px solid #1e2535; border-radius:8px; padding:0.5rem 1rem; font-size:0.8rem;">
#             <span style="color:#c9a84c; letter-spacing:1px; text-transform:uppercase; font-size:0.65rem;">Shapes</span><br>
#             <span style="color:#f0e6cc; font-weight:500;">{len(all_shapes)}</span>
#         </div>
#         <div style="background:#0d1220; border:1px solid #1e2535; border-radius:8px; padding:0.5rem 1rem; font-size:0.8rem;">
#             <span style="color:#c9a84c; letter-spacing:1px; text-transform:uppercase; font-size:0.65rem;">Cost Column</span><br>
#             <span style="color:#f0e6cc; font-weight:500;">{cost_col}</span>
#         </div>
#     </div>
#     """, unsafe_allow_html=True)

#     col_main, col_side = st.columns([3, 2], gap="large")

#     with col_main:

#         # Q1 — Output mode
#         st.markdown("""
#         <div class="q-card">
#             <div class="q-num">Question 01</div>
#             <div class="q-title">Output Mode</div>
#             <div class="q-hint">Choose how the repriced file should be generated and distributed.</div>
#         </div>
#         """, unsafe_allow_html=True)

#         team_mode = st.radio(
#             "Output Mode",
#             ["Generate individual files per team member", "Generate a single combined file"],
#             label_visibility="collapsed"
#         )

#         st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

#         # Q2 — Shape selection (only if single file)
#         if team_mode == "Generate a single combined file":
#             st.markdown("""
#             <div class="q-card">
#                 <div class="q-num">Question 02</div>
#                 <div class="q-title">Shape Selection</div>
#                 <div class="q-hint">Select one or more shapes to include in the output. Defaults to all.</div>
#             </div>
#             """, unsafe_allow_html=True)

#             selected_shapes = st.multiselect(
#                 "Select Shapes",
#                 all_shapes,
#                 default=all_shapes,
#                 label_visibility="collapsed"
#             )
#         else:
#             selected_shapes = all_shapes

#         # Q3 — Fancy colors
#         st.markdown("""
#         <div class="q-card">
#             <div class="q-num">Question 03</div>
#             <div class="q-title">Fancy Color Stones</div>
#             <div class="q-hint">Fancy colors are stones outside the D–I grading range.</div>
#         </div>
#         """, unsafe_allow_html=True)

#         fancy_option = st.radio(
#             "Fancy Color Stones",
#             ["Include all colors", "Exclude fancy colors (D–I only)"],
#             label_visibility="collapsed"
#         )

#         # Q4 — Lot filtering
#         st.markdown("""
#         <div class="q-card">
#             <div class="q-num">Question 04</div>
#             <div class="q-title">Lot Filtering</div>
#             <div class="q-hint">Flagged lots: Lot numbers starting with VP / VJ or ending with A / B.</div>
#         </div>
#         """, unsafe_allow_html=True)

#         lot_option = st.radio(
#             "Lot Filtering",
#             ["Keep all lots", "Remove flagged lots"],
#             label_visibility="collapsed"
#         )

#         # Q5 — Minimum cost
#         st.markdown("""
#         <div class="q-card">
#             <div class="q-num">Question 05</div>
#             <div class="q-title">Minimum Cost Threshold</div>
#             <div class="q-hint">Apply a floor price to remove low-value stones from the output.</div>
#         </div>
#         """, unsafe_allow_html=True)

#         six_option = st.radio(
#             "Minimum Cost Threshold",
#             ["Include all price points", "6.00 and above only"],
#             label_visibility="collapsed"
#         )

#         st.markdown("<div style='height:1.2rem'></div>", unsafe_allow_html=True)

#         ready = True
#         if team_mode == "Generate a single combined file" and not selected_shapes:
#             st.warning("Please select at least one shape before processing.")
#             ready = False

#         process_btn = st.button("✦  Process & Generate Files  ✦", disabled=not ready)

#     # -------------------------------------------------
#     # SIDE PANEL — Team Reference
#     # -------------------------------------------------
#     with col_side:
#         st.markdown("""
#         <div class="q-card" style="margin-bottom:1rem;">
#             <div class="q-num">Team Reference</div>
#             <div class="q-title">Shape Assignments</div>
#             <div class="q-hint">Files will be split automatically by these assignments in Team mode.</div>
#         </div>
#         """, unsafe_allow_html=True)

#         for member, shapes in TEAM_SHAPES.items():
#             icon = TEAM_ICONS.get(member, "◆")
#             shapes_str = " · ".join(shapes)
#             st.markdown(f"""
#             <div style="background:#0a0f1a; border:1px solid #1e2535; border-radius:10px;
#                         padding:0.9rem 1.1rem; margin-bottom:0.7rem;">
#                 <div style="font-size:0.65rem; letter-spacing:1.5px; text-transform:uppercase;
#                             color:#c9a84c; margin-bottom:0.25rem;">{icon} {member}</div>
#                 <div style="font-size:0.82rem; color:#9ca3af; font-weight:300;">{shapes_str}</div>
#             </div>
#             """, unsafe_allow_html=True)

#         # Live filter preview
#         st.markdown("""
#         <div class="q-card" style="margin-top:0.5rem;">
#             <div class="q-num">Active Settings</div>
#             <div class="q-title">Filter Summary</div>
#         </div>
#         """, unsafe_allow_html=True)

#         fancy_label = "All colors" if fancy_option == "Include all colors" else "D–I only"
#         lot_label   = "All lots"   if lot_option == "Keep all lots"        else "Flagged removed"
#         price_label = "No minimum" if six_option == "Include all price points" else "≥ 6.00"
#         mode_label  = "Team files (ZIP)" if team_mode == "Generate individual files per team member" else "Single file"

#         for k, v in [("Mode", mode_label), ("Colors", fancy_label), ("Lots", lot_label), ("Price floor", price_label)]:
#             st.markdown(f"""
#             <div style="display:flex; justify-content:space-between; align-items:center;
#                         padding:0.4rem 0; border-bottom:1px solid #1a2235;">
#                 <span style="font-size:0.78rem; color:#6b7280; letter-spacing:0.5px;">{k}</span>
#                 <span style="font-size:0.8rem; color:#d6c9b0; font-weight:500;">{v}</span>
#             </div>
#             """, unsafe_allow_html=True)

#     # -------------------------------------------------
#     # PROCESS
#     # -------------------------------------------------
#     if process_btn:

#         st.markdown("<div class='gold-divider'></div>", unsafe_allow_html=True)

#         with st.spinner("Processing your diamonds..."):

#             # ── TEAM MODE ──
#             if team_mode == "Generate individual files per team member":

#                 zip_buffer = io.BytesIO()
#                 generated  = []

#                 with zipfile.ZipFile(zip_buffer, "w") as zf:
#                     for team, team_shapes in TEAM_SHAPES.items():
#                         team_df = df.copy()
#                         shapes_for_team = [s for s in all_shapes if any(x in s for x in team_shapes)]
#                         team_df = apply_filters(team_df, shapes_for_team, fancy_option, lot_option, six_option)
#                         if len(team_df) > 0:
#                             file_data = add_formula_excel(team_df, cost_col)
#                             zf.writestr(f"{team}.xlsx", file_data.read())
#                             generated.append((team, len(team_df)))

#                 zip_buffer.seek(0)

#                 st.markdown(f"""
#                 <div class="result-banner">
#                     <div class="result-icon">📦</div>
#                     <div class="result-text">
#                         <strong>{len(generated)} team files</strong> packed into a ZIP archive —
#                         {', '.join([f"{TEAM_ICONS.get(t,'◆')} {t} ({n:,})" for t,n in generated])}
#                     </div>
#                 </div>
#                 """, unsafe_allow_html=True)

#                 st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
#                 st.download_button(
#                     "⬇  Download Team ZIP Archive",
#                     data=zip_buffer,
#                     file_name="Repricing_Team_Files.zip",
#                     mime="application/zip"
#                 )

#             # ── SINGLE FILE MODE ──
#             else:
#                 data = apply_filters(df.copy(), selected_shapes, fancy_option, lot_option, six_option)
#                 final_file = add_formula_excel(data, cost_col)

#                 st.markdown(f"""
#                 <div class="result-banner">
#                     <div class="result-icon">✅</div>
#                     <div class="result-text">
#                         Output ready — <strong>{len(data):,} stones</strong> across
#                         <strong>{len(selected_shapes)} shape(s)</strong> with formula columns appended.
#                     </div>
#                 </div>
#                 """, unsafe_allow_html=True)

#                 st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
#                 st.download_button(
#                     "⬇  Download Repriced File",
#                     data=final_file,
#                     file_name="Diamond_Repriced.xlsx",
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                 )

# else:
#     # Empty state
#     st.markdown("""
#     <div style="text-align:center; padding: 3rem 0; color:#2a3550;">
#         <div style="font-size:3rem; margin-bottom:1rem; opacity:0.4;">💎</div>
#         <div style="font-family:'Cormorant Garamond',serif; font-size:1.3rem; color:#2a3550; letter-spacing:1px;">
#             Upload a file to begin
#         </div>
#     </div>
#     """, unsafe_allow_html=True)


import streamlit as st
import pandas as pd
import io
import zipfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# -------------------------------------------------
# PAGE CONFIG
# -------------------------------------------------
st.set_page_config(
    page_title="Diamond Repricing Studio",
    page_icon="💎",
    layout="wide"
)

# -------------------------------------------------
# LUXURY CSS
# -------------------------------------------------
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@300;400;500;600&family=Jost:wght@300;400;500;600&display=swap');

/* ── Base ── */
html, body, [class*="css"] {
    font-family: 'Jost', sans-serif;
    background: #080c14;
    color: #d6c9b0;
}
.stApp { background: #080c14; }
header[data-testid="stHeader"] { background: transparent; }
section[data-testid="stSidebar"] { display: none; }

/* ── Hero ── */
.hero-wrap {
    text-align: center;
    padding: 3.2rem 0 2rem 0;
    position: relative;
}
.hero-gem {
    font-size: 2.8rem;
    margin-bottom: 0.5rem;
    display: block;
    animation: floatgem 3s ease-in-out infinite;
}
@keyframes floatgem {
    0%,100% { transform: translateY(0); }
    50%      { transform: translateY(-6px); }
}
.hero-title {
    font-family: 'Cormorant Garamond', serif;
    font-size: 3rem;
    font-weight: 400;
    letter-spacing: 2px;
    color: #f0e6cc;
    margin: 0 0 0.4rem 0;
    line-height: 1.1;
}
.hero-sub {
    font-size: 0.82rem;
    font-weight: 300;
    letter-spacing: 3px;
    text-transform: uppercase;
    color: #c9a84c;
    margin-bottom: 0;
}
.gold-line {
    width: 60px;
    height: 1px;
    background: linear-gradient(90deg, transparent, #c9a84c, transparent);
    margin: 1.2rem auto 0 auto;
}

/* ── Upload zone ── */
.upload-section {
    background: #0d1220;
    border: 1px solid #1e2535;
    border-radius: 16px;
    padding: 2rem;
    margin-bottom: 1.5rem;
    position: relative;
    overflow: hidden;
}
.upload-section::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, transparent, #c9a84c 50%, transparent);
}
.section-label {
    font-family: 'Cormorant Garamond', serif;
    font-size: 1.4rem;
    font-weight: 400;
    color: #f0e6cc;
    margin-bottom: 0.2rem;
}
.section-desc {
    font-size: 0.78rem;
    font-weight: 300;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    color: #c9a84c;
    margin-bottom: 1rem;
}

/* ── Question cards ── */
.q-card {
    background: #0d1220;
    border: 1px solid #1e2535;
    border-radius: 12px;
    padding: 1.4rem 1.6rem;
    margin-bottom: 1rem;
    transition: border-color 0.25s;
    position: relative;
}
.q-card:hover { border-color: #c9a84c44; }
.q-num {
    font-size: 0.65rem;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: #c9a84c;
    font-weight: 500;
    margin-bottom: 0.35rem;
}
.q-title {
    font-family: 'Cormorant Garamond', serif;
    font-size: 1.15rem;
    font-weight: 500;
    color: #f0e6cc;
    margin-bottom: 0.15rem;
}
.q-hint {
    font-size: 0.76rem;
    color: #6b7280;
    font-weight: 300;
    margin-bottom: 0.8rem;
}

/* ── Streamlit radio overrides ── */
div[data-testid="stRadio"] label {
    font-family: 'Jost', sans-serif !important;
    font-size: 0.9rem !important;
    color: #d6c9b0 !important;
    font-weight: 400 !important;
}
div[data-testid="stRadio"] > div { gap: 0.5rem; }

/* ── Multiselect ── */
div[data-testid="stMultiSelect"] span {
    background: #1a2235 !important;
    color: #d6c9b0 !important;
    border: 1px solid #2a3550 !important;
    border-radius: 6px !important;
}
div[data-testid="stMultiSelect"] [data-baseweb="tag"] {
    background: #1f2d45 !important;
}

/* ── File uploader ── */
div[data-testid="stFileUploader"] section {
    background: #0a0f1a !important;
    border: 1.5px dashed #2a3550 !important;
    border-radius: 12px !important;
    transition: border-color 0.2s;
}
div[data-testid="stFileUploader"] section:hover {
    border-color: #c9a84c !important;
}
div[data-testid="stFileUploader"] p { color: #6b7280 !important; }

/* ── Process button ── */
div[data-testid="stButton"] > button {
    background: linear-gradient(135deg, #c9a84c 0%, #a8833a 100%);
    color: #080c14;
    border: none;
    border-radius: 10px;
    padding: 0.85rem 2rem;
    font-family: 'Jost', sans-serif;
    font-size: 0.9rem;
    font-weight: 600;
    letter-spacing: 2px;
    text-transform: uppercase;
    width: 100%;
    cursor: pointer;
    box-shadow: 0 4px 24px rgba(201, 168, 76, 0.25);
    transition: opacity 0.2s, transform 0.15s, box-shadow 0.2s;
}
div[data-testid="stButton"] > button:hover {
    opacity: 0.92;
    transform: translateY(-2px);
    box-shadow: 0 8px 30px rgba(201, 168, 76, 0.35);
}
div[data-testid="stButton"] > button:disabled {
    background: #1e2535 !important;
    color: #3a4560 !important;
    box-shadow: none !important;
    transform: none !important;
}

/* ── Download buttons ── */
div[data-testid="stDownloadButton"] > button {
    background: #0d1220;
    color: #c9a84c;
    border: 1px solid #c9a84c55;
    border-radius: 10px;
    padding: 0.7rem 1.4rem;
    font-family: 'Jost', sans-serif;
    font-size: 0.85rem;
    font-weight: 500;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    width: 100%;
    transition: background 0.2s, border-color 0.2s;
}
div[data-testid="stDownloadButton"] > button:hover {
    background: #c9a84c15;
    border-color: #c9a84c;
}

/* ── Info / success alerts ── */
div[data-testid="stAlert"] {
    border-radius: 10px;
    font-family: 'Jost', sans-serif;
    font-size: 0.88rem;
}

/* ── Divider ── */
.gold-divider {
    height: 1px;
    background: linear-gradient(90deg, transparent, #c9a84c44, transparent);
    margin: 2rem 0;
}

/* ── Team badges ── */
.team-grid {
    display: flex; flex-wrap: wrap; gap: 0.7rem; margin-top: 0.5rem;
}
.team-badge {
    background: #111827;
    border: 1px solid #2a3550;
    border-radius: 8px;
    padding: 0.4rem 0.9rem;
    font-size: 0.78rem;
    color: #d6c9b0;
    letter-spacing: 0.5px;
}
.team-badge strong { color: #c9a84c; }

/* ── Result banner ── */
.result-banner {
    background: linear-gradient(135deg, #0d1a10, #0d1220);
    border: 1px solid #1a5c2a;
    border-radius: 12px;
    padding: 1.2rem 1.6rem;
    margin-top: 1rem;
    display: flex;
    align-items: center;
    gap: 1rem;
}
.result-icon { font-size: 1.6rem; }
.result-text { font-size: 0.9rem; color: #6fcf97; }
.result-text strong { color: #a8e6bf; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# -------------------------------------------------
# TEAM CONFIG
# -------------------------------------------------
TEAM_SHAPES = {
    "Love":    ["RBC", "ROUND", "ASSCHER", "PRINCESS"],
    "Harshali":["CUSHION MODIFIED", "CUSHION BRILLIANT", "HEART"],
    "Gautam":  ["OVAL", "EMERALD"],
    "Milan":   ["PEAR", "RADIANT"]
}

TEAM_ICONS = {"Love": "🌸", "Harshali": "💫", "Gautam": "🌿", "Milan": "✦"}

# -------------------------------------------------
# HELPERS
# -------------------------------------------------
def get_team(shape):
    s = str(shape).upper()
    for team, shapes in TEAM_SHAPES.items():
        if any(x in s for x in shapes):
            return team
    return "Others"


def add_formula_excel(df, cost_col_name, cts_col_name, lot_col_name, apply_pointer_formulas=False):
    """
    Writes df to Excel, appends formula columns, and optionally appends
    three extra value-based formula columns for U-prefix lots and
    pointer stones (< 1.00 cts, size group 0.90–0.99).

    Standard formula columns (always added):
        • Updated Price   (blank — to be filled by user)
        • Difference      = -ROUND((Cost/Cts - Updated Price) / Cost/Cts * 100, 2)

    Extra formula columns (added when apply_pointer_formulas=True):
        • Total Cost      = Cost/Cts  *  Cts          [col 1]
        • Updated Amount  = Updated Price  *  Cts      [col 2]
        • Amt Difference  = Updated Amount - Total Cost [col 3]

    The extra columns are only *populated* for rows where:
        - Lot # starts with "U"  (case-insensitive)  OR
        - Cts value is < 1.00 (pointer / size-group stone, 0.90–0.99)
    For all other rows the extra cells are left blank.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # ── Locate required columns ──
    cost_col_idx = headers.index(cost_col_name) + 1          # 1-based
    cost_letter  = get_column_letter(cost_col_idx)

    # Cts column (carats) — needed for pointer formula
    try:
        cts_col_idx  = headers.index(cts_col_name) + 1
        cts_letter   = get_column_letter(cts_col_idx)
    except ValueError:
        cts_col_idx  = None
        cts_letter   = None

    # Lot # column — needed for U-prefix check
    try:
        lot_col_idx  = headers.index(lot_col_name) + 1
        lot_letter   = get_column_letter(lot_col_idx)
    except ValueError:
        lot_col_idx  = None
        lot_letter   = None

    # ── Standard columns ──
    upd_col  = len(headers) + 1
    diff_col = len(headers) + 2
    upd_letter  = get_column_letter(upd_col)

    ws.cell(1, upd_col).value  = "Updated Price"
    ws.cell(1, diff_col).value = "Difference"

    for row in range(2, ws.max_row + 1):
        ws.cell(row, diff_col).value = (
            f"=-ROUND(({cost_letter}{row}-{upd_letter}{row})/{cost_letter}{row}*100,2)"
        )

    # ── Extra pointer / U-lot formula columns ──
    if apply_pointer_formulas and cts_col_idx and lot_col_idx:
        total_cost_col   = len(headers) + 3
        updated_amt_col  = len(headers) + 4
        amt_diff_col     = len(headers) + 5

        total_cost_letter  = get_column_letter(total_cost_col)
        updated_amt_letter = get_column_letter(updated_amt_col)

        ws.cell(1, total_cost_col).value  = "Total Cost"
        ws.cell(1, updated_amt_col).value = "Updated Amount"
        ws.cell(1, amt_diff_col).value    = "Amt Difference"

        for row in range(2, ws.max_row + 1):
            lot_cell = ws.cell(row, lot_col_idx).value
            cts_cell = ws.cell(row, cts_col_idx).value

            # Determine if this row qualifies
            lot_str   = str(lot_cell).strip().upper() if lot_cell is not None else ""
            is_u_lot  = lot_str.startswith("U")

            try:
                cts_val   = float(cts_cell)
                is_pointer = cts_val < 1.00
            except (TypeError, ValueError):
                is_pointer = False

            if is_u_lot or is_pointer:
                # Col 1: Cost/Cts * Cts  →  Total Cost
                ws.cell(row, total_cost_col).value = (
                    f"={cost_letter}{row}*{cts_letter}{row}"
                )
                # Col 2: Updated Price * Cts  →  Updated Amount
                ws.cell(row, updated_amt_col).value = (
                    f"={upd_letter}{row}*{cts_letter}{row}"
                )
                # Col 3: Updated Amount - Total Cost  →  Amt Difference
                ws.cell(row, amt_diff_col).value = (
                    f"={updated_amt_letter}{row}-{total_cost_letter}{row}"
                )
            # else: leave blank for non-qualifying rows

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final


def apply_filters(data, shapes_selected, fancy_option, lot_option, six_option):
    data = data[data[shape_col].astype(str).str.upper().isin(shapes_selected)]
    if fancy_option == "Exclude fancy colors (D–I only)":
        data = data[data[color_col].astype(str).str.upper().isin(["D","E","F","G","H","I"])]
    if lot_option == "Remove flagged lots":
        lot = data[lot_col].astype(str).str.upper()
        data = data[~(lot.str.startswith("VP") | lot.str.startswith("VJ") | lot.str.endswith("A") | lot.str.endswith("B"))]
    if six_option == "6.00 and above only":
        data = data[pd.to_numeric(data[cost_col], errors="coerce") >= 6.00]
    return data

# -------------------------------------------------
# HERO
# -------------------------------------------------
st.markdown("""
<div class="hero-wrap">
    <span class="hero-gem">💎</span>
    <h1 class="hero-title">Diamond Repricing Studio</h1>
    <p class="hero-sub">Smart Repricing · Team Distribution · Formula Ready</p>
    <div class="gold-line"></div>
</div>
""", unsafe_allow_html=True)

# -------------------------------------------------
# FILE UPLOAD
# -------------------------------------------------
st.markdown("""
<div class="upload-section">
    <div class="section-label">Source File</div>
    <div class="section-desc">Upload your master diamond inventory</div>
""", unsafe_allow_html=True)

file = st.file_uploader(
    "Upload Excel File",
    type=["xlsx"],
    label_visibility="collapsed"
)

st.markdown("</div>", unsafe_allow_html=True)

# -------------------------------------------------
# MAIN FLOW
# -------------------------------------------------
if file:
    df = pd.read_excel(file)
    df.columns = df.columns.str.strip()

    shape_col = "Shape"
    color_col = "Color"
    lot_col   = "Lot #"

    # ── Exact column names as seen in the file ──
    # Primary candidates in priority order
    COST_CANDIDATES = ["Cost / Cts.", "Cost / Cts", "Cost/Cts.", "Cost/Cts"]
    CTS_CANDIDATES  = ["Cts.", "Cts", "Carats", "Carat", "Ct"]

    stripped_cols = {c.strip(): c for c in df.columns}   # stripped → original

    cost_col = None
    for candidate in COST_CANDIDATES:
        if candidate in stripped_cols:
            cost_col = stripped_cols[candidate]
            break
    # fallback: fuzzy match
    if cost_col is None:
        for c in df.columns:
            if "cost" in c.lower() and "cts" in c.lower():
                cost_col = c
                break

    if cost_col is None:
        st.error("⚠️ Could not locate a Cost / Cts. column. Please verify your file headers.")
        st.stop()

    cts_col = None
    for candidate in CTS_CANDIDATES:
        if candidate in stripped_cols:
            cts_col = stripped_cols[candidate]
            break

    all_shapes = sorted(df[shape_col].dropna().astype(str).str.upper().unique())

    # File summary pill
    st.markdown(f"""
    <div style="display:flex; gap:1rem; margin-bottom:1.5rem; flex-wrap:wrap;">
        <div style="background:#0d1220; border:1px solid #1e2535; border-radius:8px; padding:0.5rem 1rem; font-size:0.8rem;">
            <span style="color:#c9a84c; letter-spacing:1px; text-transform:uppercase; font-size:0.65rem;">Stones</span><br>
            <span style="color:#f0e6cc; font-weight:500;">{len(df):,}</span>
        </div>
        <div style="background:#0d1220; border:1px solid #1e2535; border-radius:8px; padding:0.5rem 1rem; font-size:0.8rem;">
            <span style="color:#c9a84c; letter-spacing:1px; text-transform:uppercase; font-size:0.65rem;">Shapes</span><br>
            <span style="color:#f0e6cc; font-weight:500;">{len(all_shapes)}</span>
        </div>
        <div style="background:#0d1220; border:1px solid #1e2535; border-radius:8px; padding:0.5rem 1rem; font-size:0.8rem;">
            <span style="color:#c9a84c; letter-spacing:1px; text-transform:uppercase; font-size:0.65rem;">Cost / Cts. Column</span><br>
            <span style="color:#f0e6cc; font-weight:500;">{cost_col}</span>
        </div>
        {"" if cts_col is None else f'''
        <div style="background:#0d1220; border:1px solid #1e2535; border-radius:8px; padding:0.5rem 1rem; font-size:0.8rem;">
            <span style="color:#c9a84c; letter-spacing:1px; text-transform:uppercase; font-size:0.65rem;">Cts Column</span><br>
            <span style="color:#f0e6cc; font-weight:500;">{cts_col}</span>
        </div>'''}
    </div>
    """, unsafe_allow_html=True)

    col_main, col_side = st.columns([3, 2], gap="large")

    with col_main:

        # Q1 — Output mode
        st.markdown("""
        <div class="q-card">
            <div class="q-num">Question 01</div>
            <div class="q-title">Output Mode</div>
            <div class="q-hint">Choose how the repriced file should be generated and distributed.</div>
        </div>
        """, unsafe_allow_html=True)

        team_mode = st.radio(
            "Output Mode",
            ["Generate individual files per team member", "Generate a single combined file"],
            label_visibility="collapsed"
        )

        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

        # Q2 — Shape selection (only if single file)
        if team_mode == "Generate a single combined file":
            st.markdown("""
            <div class="q-card">
                <div class="q-num">Question 02</div>
                <div class="q-title">Shape Selection</div>
                <div class="q-hint">Select one or more shapes to include in the output. Defaults to all.</div>
            </div>
            """, unsafe_allow_html=True)

            selected_shapes = st.multiselect(
                "Select Shapes",
                all_shapes,
                default=all_shapes,
                label_visibility="collapsed"
            )
        else:
            selected_shapes = all_shapes

        # Q3 — Fancy colors
        st.markdown("""
        <div class="q-card">
            <div class="q-num">Question 03</div>
            <div class="q-title">Fancy Color Stones</div>
            <div class="q-hint">Fancy colors are stones outside the D–I grading range.</div>
        </div>
        """, unsafe_allow_html=True)

        fancy_option = st.radio(
            "Fancy Color Stones",
            ["Include all colors", "Exclude fancy colors (D–I only)"],
            label_visibility="collapsed"
        )

        # Q4 — Lot filtering
        st.markdown("""
        <div class="q-card">
            <div class="q-num">Question 04</div>
            <div class="q-title">Lot Filtering</div>
            <div class="q-hint">Flagged lots: Lot numbers starting with VP / VJ or ending with A / B.</div>
        </div>
        """, unsafe_allow_html=True)

        lot_option = st.radio(
            "Lot Filtering",
            ["Keep all lots", "Remove flagged lots"],
            label_visibility="collapsed"
        )

        # Q5 — Minimum cost
        st.markdown("""
        <div class="q-card">
            <div class="q-num">Question 05</div>
            <div class="q-title">Minimum Cost Threshold</div>
            <div class="q-hint">Apply a floor price to remove low-value stones from the output.</div>
        </div>
        """, unsafe_allow_html=True)

        six_option = st.radio(
            "Minimum Cost Threshold",
            ["Include all price points", "6.00 and above only"],
            label_visibility="collapsed"
        )

        # Q6 — Pointer / U-lot formula columns
        st.markdown("""
        <div class="q-card">
            <div class="q-num">Question 06</div>
            <div class="q-title">Pointer &amp; U-Lot Value Formulas</div>
            <div class="q-hint">
                Appends three extra formula columns — <strong>Total Cost</strong> (Cost/Cts × Cts),
                <strong>Updated Amount</strong> (Updated Price × Cts), and
                <strong>Amt Difference</strong> (Updated Amount − Total Cost) —
                populated only for stones whose Lot # starts with <strong>U</strong>
                (e.g. U51236) or whose weight is a pointer under 1.00 ct (size group 0.90–0.99).
            </div>
        </div>
        """, unsafe_allow_html=True)

        pointer_formula_option = st.radio(
            "Pointer & U-Lot Value Formulas",
            ["Do not add pointer / U-lot formulas", "Add pointer / U-lot formula columns"],
            label_visibility="collapsed"
        )

        apply_pointer_formulas = (pointer_formula_option == "Add pointer / U-lot formula columns")

        # Warn if Cts column not detected and formulas requested
        if apply_pointer_formulas and cts_col is None:
            st.warning(
                "⚠️ Could not locate a Cts. (carats) column. "
                "Pointer / U-lot formulas require a column named **Cts.** (with the dot), **Cts**, **Carats**, **Carat**, or **Ct**. "
                "Please verify your file headers and re-upload."
            )
            apply_pointer_formulas = False

        st.markdown("<div style='height:1.2rem'></div>", unsafe_allow_html=True)

        ready = True
        if team_mode == "Generate a single combined file" and not selected_shapes:
            st.warning("Please select at least one shape before processing.")
            ready = False

        process_btn = st.button("✦  Process & Generate Files  ✦", disabled=not ready)

    # -------------------------------------------------
    # SIDE PANEL — Team Reference
    # -------------------------------------------------
    with col_side:
        st.markdown("""
        <div class="q-card" style="margin-bottom:1rem;">
            <div class="q-num">Team Reference</div>
            <div class="q-title">Shape Assignments</div>
            <div class="q-hint">Files will be split automatically by these assignments in Team mode.</div>
        </div>
        """, unsafe_allow_html=True)

        for member, shapes in TEAM_SHAPES.items():
            icon = TEAM_ICONS.get(member, "◆")
            shapes_str = " · ".join(shapes)
            st.markdown(f"""
            <div style="background:#0a0f1a; border:1px solid #1e2535; border-radius:10px;
                        padding:0.9rem 1.1rem; margin-bottom:0.7rem;">
                <div style="font-size:0.65rem; letter-spacing:1.5px; text-transform:uppercase;
                            color:#c9a84c; margin-bottom:0.25rem;">{icon} {member}</div>
                <div style="font-size:0.82rem; color:#9ca3af; font-weight:300;">{shapes_str}</div>
            </div>
            """, unsafe_allow_html=True)

        # Live filter preview
        st.markdown("""
        <div class="q-card" style="margin-top:0.5rem;">
            <div class="q-num">Active Settings</div>
            <div class="q-title">Filter Summary</div>
        </div>
        """, unsafe_allow_html=True)

        fancy_label   = "All colors"      if fancy_option   == "Include all colors"               else "D–I only"
        lot_label     = "All lots"        if lot_option     == "Keep all lots"                    else "Flagged removed"
        price_label   = "No minimum"      if six_option     == "Include all price points"         else "≥ 6.00"
        mode_label    = "Team files (ZIP)" if team_mode     == "Generate individual files per team member" else "Single file"
        pointer_label = "Yes"             if apply_pointer_formulas                               else "No"

        for k, v in [
            ("Mode",          mode_label),
            ("Colors",        fancy_label),
            ("Lots",          lot_label),
            ("Price floor",   price_label),
            ("U-lot/Pointer", pointer_label),
        ]:
            st.markdown(f"""
            <div style="display:flex; justify-content:space-between; align-items:center;
                        padding:0.4rem 0; border-bottom:1px solid #1a2235;">
                <span style="font-size:0.78rem; color:#6b7280; letter-spacing:0.5px;">{k}</span>
                <span style="font-size:0.8rem; color:#d6c9b0; font-weight:500;">{v}</span>
            </div>
            """, unsafe_allow_html=True)

    # -------------------------------------------------
    # PROCESS
    # -------------------------------------------------
    if process_btn:

        st.markdown("<div class='gold-divider'></div>", unsafe_allow_html=True)

        with st.spinner("Processing your diamonds..."):

            # ── TEAM MODE ──
            if team_mode == "Generate individual files per team member":

                zip_buffer = io.BytesIO()
                generated  = []

                with zipfile.ZipFile(zip_buffer, "w") as zf:
                    for team, team_shapes in TEAM_SHAPES.items():
                        team_df = df.copy()
                        shapes_for_team = [s for s in all_shapes if any(x in s for x in team_shapes)]
                        team_df = apply_filters(team_df, shapes_for_team, fancy_option, lot_option, six_option)
                        if len(team_df) > 0:
                            file_data = add_formula_excel(
                                team_df, cost_col,
                                cts_col_name=cts_col if cts_col else "",
                                lot_col_name=lot_col,
                                apply_pointer_formulas=apply_pointer_formulas
                            )
                            zf.writestr(f"{team}.xlsx", file_data.read())
                            generated.append((team, len(team_df)))

                zip_buffer.seek(0)

                st.markdown(f"""
                <div class="result-banner">
                    <div class="result-icon">📦</div>
                    <div class="result-text">
                        <strong>{len(generated)} team files</strong> packed into a ZIP archive —
                        {', '.join([f"{TEAM_ICONS.get(t,'◆')} {t} ({n:,})" for t,n in generated])}
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
                st.download_button(
                    "⬇  Download Team ZIP Archive",
                    data=zip_buffer,
                    file_name="Repricing_Team_Files.zip",
                    mime="application/zip"
                )

            # ── SINGLE FILE MODE ──
            else:
                data = apply_filters(df.copy(), selected_shapes, fancy_option, lot_option, six_option)
                final_file = add_formula_excel(
                    data, cost_col,
                    cts_col_name=cts_col if cts_col else "",
                    lot_col_name=lot_col,
                    apply_pointer_formulas=apply_pointer_formulas
                )

                st.markdown(f"""
                <div class="result-banner">
                    <div class="result-icon">✅</div>
                    <div class="result-text">
                        Output ready — <strong>{len(data):,} stones</strong> across
                        <strong>{len(selected_shapes)} shape(s)</strong> with formula columns appended.
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
                st.download_button(
                    "⬇  Download Repriced File",
                    data=final_file,
                    file_name="Diamond_Repriced.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

else:
    # Empty state
    st.markdown("""
    <div style="text-align:center; padding: 3rem 0; color:#2a3550;">
        <div style="font-size:3rem; margin-bottom:1rem; opacity:0.4;">💎</div>
        <div style="font-family:'Cormorant Garamond',serif; font-size:1.3rem; color:#2a3550; letter-spacing:1px;">
            Upload a file to begin
        </div>
    </div>
    """, unsafe_allow_html=True)